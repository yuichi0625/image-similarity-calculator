import csv
import os

import numpy as np
from natsort import natsorted, ns
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.fonts import Font
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

WIDTH = 100


def output_csv(sim_scores, sim_src_paths, csv_path):
    with open(csv_path, 'w') as f:
        writer = csv.writer(f, lineterminator='\n')
        writer.writerow(['file1', 'file2', 'score'])
        for score, (path1, path2) in zip(sim_scores, natsorted(sim_src_paths, alg=ns.IGNORECASE)):
            file1 = os.path.basename(path1)
            file2 = os.path.basename(path2)
            writer.writerow([file1, file2, score])


def output_excel(sim_scores, sim_src_paths, excel_path):
    wb = Workbook()
    ws = wb.active
    ws.title = os.path.basename(excel_path).split('.')[0]

    # read images and resize them for displaying
    paths = natsorted(np.unique(sim_src_paths), alg=ns.IGNORECASE)
    imgs = [PILImage.open(path) for path in paths]
    resized_imgs = [img.resize((WIDTH, int(img.size[1] / img.size[0] * WIDTH))) for img in imgs]
    max_height = max(resized_img.size[1] for resized_img in resized_imgs)

    for i, (path, resized_img) in enumerate(zip(paths, resized_imgs), 2):
        filename = os.path.basename(path).split('.')[0]

        # place images on the 1st row
        cell = ws.cell(row=1, column=i)
        ws[cell.coordinate] = filename
        ws[cell.coordinate].alignment = Alignment(horizontal='center', vertical='top')
        ws.row_dimensions[1].height = 14.5

        cell = ws.cell(row=2, column=i)
        ws.row_dimensions[2].height = max_height * 0.78
        ws.column_dimensions[cell.column_letter].width = WIDTH * 0.14
        img = Image(resized_img)
        img.format = 'jpg'
        img.anchor = cell.coordinate
        ws.add_image(img)

        # place images on the 1st column
        cell = ws.cell(row=i*2-1, column=1)
        ws[cell.coordinate] = filename
        ws[cell.coordinate].alignment = Alignment(horizontal='center', vertical='top')
        ws.row_dimensions[i*2-1].height = 14.5

        cell = ws.cell(row=i*2, column=1)
        ws.row_dimensions[i*2].height = resized_img.size[1] * 0.78
        ws.column_dimensions[cell.column_letter].width = WIDTH * 0.14
        img = Image(resized_img)
        img.format = 'jpg'
        img.anchor = cell.coordinate
        ws.add_image(img)

    # place similarity scores on each cell
    for score, (path1, path2) in zip(sim_scores, sim_src_paths):
        idx1 = paths.index(path1)
        idx2 = paths.index(path2)
        for row, col in [(idx1, idx2), (idx2, idx1)]:
            cell = ws.cell(row=row*2+3, column=col+2)
            hex_score = hex(int((1 - score) * 255))[2:]
            hex_score = '0' + hex_score if len(hex_score) == 1 else hex_score
            ws[cell.coordinate].fill = PatternFill(patternType='solid', fgColor=f'ff{hex_score}{hex_score}')
            ws[cell.coordinate] = f'{score:.2f}'
            ws[cell.coordinate].font = Font(size=16)
            ws[cell.coordinate].alignment = Alignment(horizontal='center', vertical='center')

    # set border and merge cells
    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)
    for row in range(1, len(paths) + 2):
        for col in range(1, len(paths) + 2):
            cell_upper = ws.cell(row=row*2-1, column=col)
            cell_lower = ws.cell(row=row*2, column=col)
            ws[cell_upper.coordinate].border = border
            ws.merge_cells(f'{cell_upper.coordinate}:{cell_lower.coordinate}')

    wb.save(excel_path)
    wb.close()
